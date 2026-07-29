#!/usr/bin/env python3
"""
AI Readiness Diagnostic Report Generator
Builds a task DAG from interview results and generates an Excel diagnostic report.

Usage:
    python3 generate_report.py --input tasks.json --output report.xlsx
    python3 generate_report.py --demo  # smoke-test with sample data
"""

import argparse
import json
import sys
import os
from datetime import datetime

import networkx as nx
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.font_manager as fm
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ─────────────────────────────────────────────
# Font setup (CJK-capable, so non-Latin task names still render)
# ─────────────────────────────────────────────
def setup_japanese_font():
    """Find and configure an available CJK-capable font."""
    candidates = [
        "Noto Sans CJK JP", "Noto Sans JP", "IPAGothic", "IPAPGothic",
        "VL Gothic", "TakaoGothic", "DejaVu Sans"
    ]
    available = {f.name for f in fm.fontManager.ttflist}
    for font in candidates:
        if font in available:
            plt.rcParams["font.family"] = font
            return font
    plt.rcParams["font.family"] = "DejaVu Sans"
    return "DejaVu Sans"

# ─────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────
# Task JSON schema:
# {
#   "business_name": "Monthly close",
#   "company": "Sample Corp",
#   "interviewer_notes": "...",
#   "tasks": [
#     {
#       "id": "T01",
#       "name": "Collect sales data",
#       "description": "Collect sales data from each department in Excel",
#       "input": "Each department's Excel file",
#       "output": "Consolidated sales data",
#       "owner": "Finance",
#       "duration_hours": 4,
#       "frequency": "Monthly",
#       "is_confidential": false,
#       "requires_human_approval": false,
#       "ai_fit": "High",           # High/Medium/Low
#       "ai_fit_reason": "Repetitive data aggregation",
#       "ai_role": "Automatic aggregation and format normalisation",
#       "dependencies": []        # list of predecessor task IDs
#     },
#     ...
#   ]
# }

DEMO_DATA = {
    "business_name": "Monthly close",
    "company": "Sample Corp",
    "interviewer_notes": "Takes 4-5 days at each month end. Data collection and re-keying are the biggest bottleneck.",
    "tasks": [
        {
            "id": "T01",
            "name": "Collect sales data",
            "description": "Collect sales data from each department in Excel",
            "input": "Each department's Excel file (email attachment)",
            "output": "Pre-consolidation sales data",
            "owner": "Finance",
            "duration_hours": 4,
            "frequency": "Monthly",
            "is_confidential": False,
            "requires_human_approval": False,
            "ai_fit": "High",
            "ai_fit_reason": "Repetitive collection in a fixed format. Email attachments can be ingested automatically",
            "ai_role": "Auto-sorting email, collecting files, normalising formats",
            "dependencies": []
        },
        {
            "id": "T02",
            "name": "Transcribe and consolidate data",
            "description": "Transcribe and consolidate the collected Excel files into one master sheet",
            "input": "Each department's Excel file",
            "output": "Consolidated sales master sheet",
            "owner": "Finance",
            "duration_hours": 6,
            "frequency": "Monthly",
            "is_confidential": False,
            "requires_human_approval": False,
            "ai_fit": "High",
            "ai_fit_reason": "Rule-based transcription. Error detection can be automated too",
            "ai_role": "Auto-transcription, duplicate checking, aggregation",
            "dependencies": ["T01"]
        },
        {
            "id": "T03",
            "name": "Enter journal entries",
            "description": "Enter journal entries into the accounting system from the consolidated data",
            "input": "Consolidated sales master sheet",
            "output": "Journal entries recorded in the accounting system",
            "owner": "Finance",
            "duration_hours": 5,
            "frequency": "Monthly",
            "is_confidential": True,
            "requires_human_approval": True,
            "ai_fit": "Medium",
            "ai_fit_reason": "Entry patterns can be turned into rules, but exceptions and the final check need a human",
            "ai_role": "Drafting entries and pattern matching (final approval by a human)",
            "dependencies": ["T02"]
        },
        {
            "id": "T04",
            "name": "Variance analysis",
            "description": "Analyse variances against last month and budget, and write commentary",
            "input": "Trial balance from the accounting system",
            "output": "Variance commentary",
            "owner": "CFO / Head of Finance",
            "duration_hours": 3,
            "frequency": "Monthly",
            "is_confidential": True,
            "requires_human_approval": False,
            "ai_fit": "Medium",
            "ai_fit_reason": "AI can compare figures and draft commentary. Interpretation involving management judgement stays with a human",
            "ai_role": "Calculating variances and drafting commentary",
            "dependencies": ["T03"]
        },
        {
            "id": "T05",
            "name": "Prepare the management report",
            "description": "Prepare the monthly management report for the board",
            "input": "Variance commentary and trial balance",
            "output": "Monthly management report (PowerPoint)",
            "owner": "CFO",
            "duration_hours": 4,
            "frequency": "Monthly",
            "is_confidential": True,
            "requires_human_approval": True,
            "ai_fit": "Medium",
            "ai_fit_reason": "AI can structure slides and insert figures. The management message is written by a human",
            "ai_role": "Auto-inserting figures into the template and proposing a slide structure",
            "dependencies": ["T04"]
        },
        {
            "id": "T06",
            "name": "Prepare audit materials",
            "description": "Organise and submit supporting documents to the external auditor",
            "input": "Accounting data and supporting documents",
            "output": "A complete set of audit files",
            "owner": "Finance",
            "duration_hours": 8,
            "frequency": "Quarterly",
            "is_confidential": True,
            "requires_human_approval": True,
            "ai_fit": "Low",
            "ai_fit_reason": "High legal and compliance significance; the final decision must always be made by a human",
            "ai_role": "Only building a document checklist and helping organise files",
            "dependencies": ["T03"]
        },
        {
            "id": "T07",
            "name": "Process invoices",
            "description": "Receive supplier invoices, reconcile them, and register payment",
            "input": "Invoices (PDF and paper)",
            "output": "Payment-registered data",
            "owner": "Finance",
            "duration_hours": 5,
            "frequency": "Monthly",
            "is_confidential": False,
            "requires_human_approval": True,
            "ai_fit": "High",
            "ai_fit_reason": "Extraction by OCR and reconciliation lend themselves well to automation",
            "ai_role": "Reading by OCR, automatic reconciliation against purchase orders, drafting the payment record",
            "dependencies": []
        }
    ]
}

# ─────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────
COLOR = {
    "high":     {"hex": "C6EFCE", "font": "276221", "label": "High"},
    "medium":   {"hex": "FFEB9C", "font": "9C5700", "label": "Medium"},
    "low":      {"hex": "FFC7CE", "font": "9C0006", "label": "Low"},
    "header":   {"hex": "1F3864", "font": "FFFFFF"},
    "subheader":{"hex": "2F5496", "font": "FFFFFF"},
    "section":  {"hex": "D6E4F0", "font": "1F3864"},
    "alt_row":  {"hex": "F2F7FB", "font": "000000"},
    "border":   "B8CCE4",
    "dag_high": "#2ECC71",
    "dag_med":  "#F39C12",
    "dag_low":  "#E74C3C",
    "dag_edge": "#7F8C8D",
}

def fit_color(ai_fit: str) -> dict:
    mapping = {"High": COLOR["high"], "Medium": COLOR["medium"], "Low": COLOR["low"]}
    return mapping.get(ai_fit, COLOR["medium"])

# ─────────────────────────────────────────────
# DAG construction
# ─────────────────────────────────────────────
def build_dag(tasks: list) -> nx.DiGraph:
    G = nx.DiGraph()
    for t in tasks:
        G.add_node(t["id"], **t)
    for t in tasks:
        for dep in t.get("dependencies", []):
            G.add_edge(dep, t["id"])
    if not nx.is_directed_acyclic_graph(G):
        raise ValueError("A cycle was detected in the task dependencies. Please review them.")
    return G

def compute_dag_metrics(G: nx.DiGraph, tasks: list) -> list:
    """Attach DAG metrics to each task (topological order, parallel group, critical path)."""
    task_map = {t["id"]: t for t in tasks}

    # topological sort
    topo_order = list(nx.topological_sort(G))

    # compute the earliest start time (EST)
    est = {}
    for node in topo_order:
        preds = list(G.predecessors(node))
        if not preds:
            est[node] = 0
        else:
            est[node] = max(
                est[p] + task_map[p]["duration_hours"] for p in preds
            )

    # parallel groups (tasks sharing an EST can run in parallel)
    groups = {}
    for node in topo_order:
        g = est[node]
        groups.setdefault(g, []).append(node)

    parallel_group = {}
    for g_idx, (_, nodes) in enumerate(sorted(groups.items()), 1):
        for node in nodes:
            parallel_group[node] = g_idx

    # identify the critical path
    critical_path = set(nx.dag_longest_path(G))

    # attach the metrics to each task
    enriched = []
    for t in tasks:
        tid = t["id"]
        enriched.append({
            **t,
            "topo_order":      topo_order.index(tid) + 1,
            "est_hours":       est[tid],
            "parallel_group":  parallel_group[tid],
            "is_critical_path": tid in critical_path,
        })
    return enriched

# ─────────────────────────────────────────────
# DAG visualisation (PNG output)
# ─────────────────────────────────────────────
def hierarchical_layout(G: nx.DiGraph) -> dict:
    """Manually compute a left-to-right layered layout from the topological order."""
    topo = list(nx.topological_sort(G))
    # decide each node's layer by longest path
    layer = {}
    for node in topo:
        preds = list(G.predecessors(node))
        layer[node] = max((layer[p] + 1 for p in preds), default=0)

    # lay out nodes in the same layer along y
    from collections import defaultdict
    layer_nodes = defaultdict(list)
    for node, l in layer.items():
        layer_nodes[l].append(node)

    pos = {}
    x_gap = 4.5
    y_gap = 2.5
    for l, nodes in sorted(layer_nodes.items()):
        n = len(nodes)
        for i, node in enumerate(nodes):
            x = l * x_gap
            y = (i - (n - 1) / 2.0) * y_gap
            pos[node] = (x, -y)  # flip y so it reads top to bottom
    return pos

def render_dag(G: nx.DiGraph, tasks: list, output_path: str, font_name: str):
    task_map = {t["id"]: t for t in tasks}
    fig, ax = plt.subplots(figsize=(18, 10))
    ax.set_facecolor("#F8FBFF")
    fig.patch.set_facecolor("#F8FBFF")

    # layout: layered, left to right
    pos = hierarchical_layout(G)

    # node colours
    node_colors = []
    for node in G.nodes():
        fit = task_map[node]["ai_fit"]
        if fit == "High":
            node_colors.append(COLOR["dag_high"])
        elif fit == "Medium":
            node_colors.append(COLOR["dag_med"])
        else:
            node_colors.append(COLOR["dag_low"])

    # emphasise critical-path edges
    critical_path_nodes = set(nx.dag_longest_path(G))
    edge_colors = []
    edge_widths = []
    for u, v in G.edges():
        if u in critical_path_nodes and v in critical_path_nodes:
            edge_colors.append("#C0392B")
            edge_widths.append(3.0)
        else:
            edge_colors.append(COLOR["dag_edge"])
            edge_widths.append(1.5)

    nx.draw_networkx_edges(
        G, pos, ax=ax,
        edge_color=edge_colors, width=edge_widths,
        arrows=True, arrowsize=20,
        connectionstyle="arc3,rad=0.05"
    )
    nx.draw_networkx_nodes(
        G, pos, ax=ax,
        node_color=node_colors, node_size=3600,
        alpha=0.95, linewidths=2.5, edgecolors="#2C3E50"
    )

    # labels (ID + name)
    labels = {}
    for node in G.nodes():
        name = task_map[node]["name"]
        short = name if len(name) <= 7 else name[:6] + "…"
        labels[node] = f"{node}\n{short}"
    nx.draw_networkx_labels(
        G, pos, labels, ax=ax,
        font_size=10, font_family=font_name, font_color="white", font_weight="bold"
    )

    # legend
    legend_handles = [
        mpatches.Patch(color=COLOR["dag_high"], label="AI fit: High"),
        mpatches.Patch(color=COLOR["dag_med"],  label="AI fit: Medium"),
        mpatches.Patch(color=COLOR["dag_low"],  label="AI fit: Low"),
        mpatches.Patch(color="#C0392B",         label="Critical path"),
    ]
    ax.legend(handles=legend_handles, loc="upper left", fontsize=10,
              prop={"family": font_name, "size": 10})

    ax.set_title("Task dependency DAG (AI fit mapping)",
                 fontsize=14, fontfamily=font_name, pad=15, fontweight="bold")
    ax.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()

# ─────────────────────────────────────────────
# Excel report generation
# ─────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, value, bg=None, font_color=None, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    bg = bg or COLOR["header"]["hex"]
    fc = font_color or COLOR["header"]["font"]
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fc, size=size, name="Meiryo UI")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell

def data_cell(ws, row, col, value, bg=None, bold=False, align="left", wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10, name="Meiryo UI")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin_border()
    return cell

def generate_excel(data: dict, enriched_tasks: list, dag_image_path: str, output_path: str):
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 55

    # title block
    ws1.merge_cells("A1:B1")
    c = ws1.cell(row=1, column=1, value="AI readiness diagnostic report")
    c.fill = PatternFill("solid", fgColor=COLOR["header"]["hex"])
    c.font = Font(bold=True, color=COLOR["header"]["font"], size=16, name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    meta = [
        ("Business area",   data.get("business_name", "")),
        ("Company",     data.get("company", "")),
        ("Date",     datetime.now().strftime("%Y-%m-%d")),
        ("Interview notes", data.get("interviewer_notes", "")),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws1.row_dimensions[i].height = 28
        data_cell(ws1, i, 1, k, bg=COLOR["section"]["hex"], bold=True, align="center")
        data_cell(ws1, i, 2, v)

    # totals
    total = len(enriched_tasks)
    high  = sum(1 for t in enriched_tasks if t["ai_fit"] == "High")
    med   = sum(1 for t in enriched_tasks if t["ai_fit"] == "Medium")
    low   = sum(1 for t in enriched_tasks if t["ai_fit"] == "Low")
    total_h = sum(t["duration_hours"] for t in enriched_tasks)
    auto_h  = sum(t["duration_hours"] for t in enriched_tasks if t["ai_fit"] == "High")
    assist_h= sum(t["duration_hours"] for t in enriched_tasks if t["ai_fit"] == "Medium")

    ws1.row_dimensions[7].height = 30
    ws1.merge_cells("A7:B7")
    header_style(ws1, 7, 1, "Totals", bg=COLOR["subheader"]["hex"], size=12)

    summary_rows = [
        ("Total tasks",          f"{total} tasks"),
        ("AI fit: High",        f"{high} tasks (automation or a large efficiency gain is realistic)"),
        ("AI fit: Medium",        f"{med} tasks (AI assistance and draft generation are effective)"),
        ("AI fit: Low",        f"{low} tasks (human judgement and oversight required)"),
        ("Total monthly effort",          f"{total_h:.1f} hours"),
        ("Effort AI could automate",  f"{auto_h:.1f} hours ({auto_h/total_h*100:.0f}%）"),
        ("Effort AI could assist",    f"{assist_h:.1f} hours ({assist_h/total_h*100:.0f}%）"),
    ]
    for i, (k, v) in enumerate(summary_rows, start=8):
        ws1.row_dimensions[i].height = 24
        data_cell(ws1, i, 1, k, bg=COLOR["alt_row"]["hex"], bold=True, align="center")
        data_cell(ws1, i, 2, v)

    # --- Sheet 2: Task list ---
    ws2 = wb.create_sheet("Task list")
    ws2.sheet_view.showGridLines = False

    col_defs = [
        ("ID",           8),
        ("Task name",     22),
        ("Description",         38),
        ("Input",         22),
        ("Output",         22),
        ("Owner",       14),
        ("Effort (h)",       9),
        ("Frequency",          9),
        ("Confidential",      9),
        ("Approval needed",        9),
        ("AI fit",     10),
        ("Why",   38),
        ("What AI can do",     38),
        ("Depends on",   14),
        ("Order",      9),
        ("Parallel group", 12),
        ("Critical", 12),
        ("Earliest start (h)",  12),
    ]
    for ci, (title, width) in enumerate(col_defs, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = width
        header_style(ws2, 1, ci, title)
    ws2.row_dimensions[1].height = 32
    ws2.freeze_panes = "A2"

    for ri, t in enumerate(enriched_tasks, start=2):
        ws2.row_dimensions[ri].height = 52
        row_bg = None if ri % 2 == 0 else COLOR["alt_row"]["hex"]
        fc = fit_color(t["ai_fit"])

        data_cell(ws2, ri, 1,  t["id"],              bg=row_bg, bold=True, align="center")
        data_cell(ws2, ri, 2,  t["name"],             bg=row_bg, bold=True)
        data_cell(ws2, ri, 3,  t["description"],      bg=row_bg)
        data_cell(ws2, ri, 4,  t["input"],            bg=row_bg)
        data_cell(ws2, ri, 5,  t["output"],           bg=row_bg)
        data_cell(ws2, ri, 6,  t["owner"],            bg=row_bg, align="center")
        data_cell(ws2, ri, 7,  t["duration_hours"],   bg=row_bg, align="center")
        data_cell(ws2, ri, 8,  t["frequency"],        bg=row_bg, align="center")
        data_cell(ws2, ri, 9,  "●" if t["is_confidential"] else "○",
                  bg=row_bg, align="center",
                  bold=t["is_confidential"])
        data_cell(ws2, ri, 10, "●" if t["requires_human_approval"] else "○",
                  bg=row_bg, align="center",
                  bold=t["requires_human_approval"])

        # AI-fit cell (colour coded)
        fit_cell = ws2.cell(row=ri, column=11, value=t["ai_fit"])
        fit_cell.fill = PatternFill("solid", fgColor=fc["hex"])
        fit_cell.font = Font(bold=True, color=fc["font"], size=11, name="Meiryo UI")
        fit_cell.alignment = Alignment(horizontal="center", vertical="center")
        fit_cell.border = thin_border()

        data_cell(ws2, ri, 12, t["ai_fit_reason"],    bg=row_bg)
        data_cell(ws2, ri, 13, t["ai_role"],          bg=row_bg)
        deps = ", ".join(t.get("dependencies", [])) or "None"
        data_cell(ws2, ri, 14, deps,                  bg=row_bg, align="center")
        data_cell(ws2, ri, 15, t["topo_order"],       bg=row_bg, align="center")
        data_cell(ws2, ri, 16, t["parallel_group"],   bg=row_bg, align="center")

        cp_cell = ws2.cell(row=ri, column=17,
                           value="* Critical" if t["is_critical_path"] else "")
        cp_bg = "FFE0E0" if t["is_critical_path"] else (row_bg or "FFFFFF")
        cp_cell.fill = PatternFill("solid", fgColor=cp_bg)
        cp_cell.font = Font(bold=t["is_critical_path"], color="C00000" if t["is_critical_path"] else "000000",
                            size=10, name="Meiryo UI")
        cp_cell.alignment = Alignment(horizontal="center", vertical="center")
        cp_cell.border = thin_border()

        data_cell(ws2, ri, 18, t["est_hours"],        bg=row_bg, align="center")

    # --- Sheet 3: DAG visual ---
    ws3 = wb.create_sheet("Dependency DAG")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:N1")
    c = ws3.cell(row=1, column=1, value="Task dependency DAG (AI fit mapping)")
    c.fill = PatternFill("solid", fgColor=COLOR["header"]["hex"])
    c.font = Font(bold=True, color=COLOR["header"]["font"], size=14, name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 36

    if os.path.exists(dag_image_path):
        img = XLImage(dag_image_path)
        img.width  = 900
        img.height = 506
        ws3.add_image(img, "A3")
        for r in range(3, 40):
            ws3.row_dimensions[r].height = 15

    # --- Sheet 4: Roadmap ---
    ws4 = wb.create_sheet("Adoption roadmap")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["C"].width = 45
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 30

    ws4.merge_cells("A1:E1")
    c = ws4.cell(row=1, column=1, value="AI adoption roadmap (recommended phases)")
    c.fill = PatternFill("solid", fgColor=COLOR["header"]["hex"])
    c.font = Font(bold=True, color=COLOR["header"]["font"], size=14, name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 36

    for ci, (title, _) in enumerate([
        ("Phase", 16), ("Tasks", 24), ("What AI can do", 45),
        ("Expected benefit", 30), ("Watch out for", 30)
    ], start=1):
        header_style(ws4, 2, ci, title, bg=COLOR["subheader"]["hex"])
    ws4.row_dimensions[2].height = 28

    # Phase 1: prioritise tasks with AI fit 'High'
    high_tasks = [t for t in enriched_tasks if t["ai_fit"] == "High"]
    med_tasks  = [t for t in enriched_tasks if t["ai_fit"] == "Medium"]
    low_tasks  = [t for t in enriched_tasks if t["ai_fit"] == "Low"]

    phases = [
        (
            "Phase 1\n(0-3 months)",
            "\n".join(f"{t['id']}: {t['name']}" for t in high_tasks) or "—",
            "\n".join(t["ai_role"] for t in high_tasks) or "—",
            f"Cuts up to {sum(t['duration_hours'] for t in high_tasks):.0f}h of monthly workload",
            "Standardising the format is a precondition. Pilot with one task first"
        ),
        (
            "Phase 2\n(3-6 months)",
            "\n".join(f"{t['id']}: {t['name']}" for t in med_tasks) or "—",
            "\n".join(t["ai_role"] for t in med_tasks) or "—",
            f"Better quality via AI assistance; around {sum(t['duration_hours'] for t in med_tasks):.0f}h of workload saved",
            "A human must give final approval. Design a review process for AI output"
        ),
        (
            "Phase 3\n(6 months+)",
            "\n".join(f"{t['id']}: {t['name']}" for t in low_tasks) or "—",
            "\n".join(t["ai_role"] for t in low_tasks) or "—",
            "Keep AI to a supporting role and minimise the cost of human oversight",
            "Involves confidential information and legal judgement, so AI is limited to a supporting tool. No external transmission"
        ),
    ]

    for ri, (phase, tasks_str, roles, effect, caution) in enumerate(phases, start=3):
        ws4.row_dimensions[ri].height = 80
        bg = COLOR["alt_row"]["hex"] if ri % 2 == 0 else None
        data_cell(ws4, ri, 1, phase,     bg=COLOR["section"]["hex"], bold=True, align="center")
        data_cell(ws4, ri, 2, tasks_str, bg=bg)
        data_cell(ws4, ri, 3, roles,     bg=bg)
        data_cell(ws4, ri, 4, effect,    bg=bg)
        data_cell(ws4, ri, 5, caution,   bg=bg)

    # footnote row
    ws4.row_dimensions[6].height = 20
    ws4.merge_cells("A7:E7")
    note = ws4.cell(row=7, column=1,
                    value="Note: for work involving confidential or personal information, always put access restrictions in place and ensure the final decision is made by a human (the Human-in-the-loop principle)")
    note.font = Font(italic=True, color="C00000", size=9, name="Meiryo UI")
    note.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(output_path)
    print(f"[OK] Excel report generated: {output_path}")

# ─────────────────────────────────────────────
# main entry point
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="AI Readiness Diagnostic Report Generator")
    parser.add_argument("--input",  help="path to the task JSON file")
    parser.add_argument("--output", default="ai-readiness-report.xlsx", help="output Excel filename")
    parser.add_argument("--demo",   action="store_true", help="run with sample data")
    args = parser.parse_args()

    if args.demo:
        data = DEMO_DATA
    elif args.input:
        with open(args.input, encoding="utf-8") as f:
            data = json.load(f)
    else:
        print("Please specify --input or --demo.")
        sys.exit(1)

    print("[1/4] Configuring font...")
    font_name = setup_japanese_font()
    print(f"      Font: {font_name}")

    print("[2/4] Building the DAG...")
    G = build_dag(data["tasks"])
    enriched = compute_dag_metrics(G, data["tasks"])
    print(f"      Tasks: {len(enriched)}, edges: {G.number_of_edges()}")

    print("[3/4] Rendering the DAG...")
    dag_img = args.output.replace(".xlsx", "_dag.png")
    render_dag(G, data["tasks"], dag_img, font_name)
    print(f"      DAG image: {dag_img}")

    print("[4/4] Generating the Excel report...")
    generate_excel(data, enriched, dag_img, args.output)

    print("\nDone.")
    print(f"  Excel: {args.output}")
    print(f"  DAG:   {dag_img}")

if __name__ == "__main__":
    main()
