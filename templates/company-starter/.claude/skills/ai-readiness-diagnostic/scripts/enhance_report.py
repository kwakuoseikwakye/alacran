#!/usr/bin/env python3
"""
AI Readiness Diagnostic Report Enhancer
Enhances the Excel report produced by generate_report.py in the following ways:
  1. Rewording by audience (--audience executive switches to plain language for executives)
  2. Swapping in a custom DAG image (--custom-dag-image, for an explanatory diagram)
  3. Adding a work-block mapping table (--block-mapping, linking diagram blocks to task IDs)

Usage:
    python3 enhance_report.py \
        --input "ai-readiness-report.xlsx" \
        --output "ai-readiness-report_final.xlsx" \
        --audience executive \
        --custom-dag-image dag.png \
        --block-mapping mapping.json
"""

import argparse
import json
import os
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# the skill's references directory
SKILL_DIR = Path(__file__).resolve().parent.parent
DEFAULT_TERMINOLOGY = SKILL_DIR / "references" / "terminology_executive.json"


# ─────────────────────────────────────────────
# load the terminology dictionary
# ─────────────────────────────────────────────
def load_terminology(path: Path) -> dict:
    """Load the dictionary of plain expressions for executives."""
    if not path.exists():
        print(f"[!] Terminology dictionary not found: {path}", file=sys.stderr)
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────
# Sheet 1: plain-language the summary
# ─────────────────────────────────────────────
def enhance_summary_sheet(ws, terminology: dict):
    """Replace the summary sheet labels with plain expressions."""
    label_map = terminology.get("summary_labels", {})
    if not label_map:
        return

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for old, new in label_map.items():
                    if cell.value.startswith(old):
                        cell.value = cell.value.replace(old, new)


# ─────────────────────────────────────────────
# Sheet 2: plain-language the task list's column names and fit labels
# ─────────────────────────────────────────────
def enhance_task_sheet(ws, terminology: dict):
    """Plain-language the task sheet's column names, fit labels and critical-path marker."""
    header_map = terminology.get("task_sheet_headers", {})
    fit_label_map = terminology.get("ai_fit_labels", {})
    critical_label = terminology.get("critical_path_label")

    # rewrite the header row's column names
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=ci)
        if cell.value in header_map:
            cell.value = header_map[cell.value]

    # locate the AI-fit column
    fit_col = None
    critical_col = None
    for ci in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=ci).value
        if h in ("How easily AI can take it on", "AI fit"):
            fit_col = ci
        if h in ("Impact on the whole", "Critical"):
            critical_col = ci

    # turn High/Medium/Low into plain labels
    if fit_col and fit_label_map:
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=fit_col)
            if cell.value in fit_label_map:
                cell.value = fit_label_map[cell.value]

    # change the critical-path marker
    if critical_col and critical_label:
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=critical_col)
            if cell.value == "* Critical":
                cell.value = critical_label


# ─────────────────────────────────────────────
# Sheet 3: swap the DAG sheet + add the mapping table
# ─────────────────────────────────────────────
def replace_dag_image(ws, custom_dag_path: str, block_mapping: list,
                     subtitle: str = None):
    """Swap the image on the DAG sheet and add the work-block mapping table."""
    # remove the existing image
    ws._images = []

    # add a subtitle (optional)
    if subtitle:
        ws.merge_cells("A2:N2")
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = Font(italic=True, color="595959", size=10, name="Meiryo UI")
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 24

    # insert the custom DAG image (preserving aspect ratio)
    if custom_dag_path and os.path.exists(custom_dag_path):
        from PIL import Image as PILImage
        with PILImage.open(custom_dag_path) as pil_img:
            aspect = pil_img.size[0] / pil_img.size[1]

        img = XLImage(custom_dag_path)
        img.width = 900
        img.height = int(900 / aspect)
        ws.add_image(img, "A4")

    # add the work-block mapping table
    if block_mapping:
        caption_row = 41
        ws.merge_cells(f"A{caption_row}:N{caption_row}")
        cap_title = ws.cell(row=caption_row, column=1,
                            value="-- How the blocks in the diagram map to the work in the task list --")
        cap_title.font = Font(bold=True, color="1F3864", size=11, name="Meiryo UI")
        cap_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[caption_row].height = 28

        start_row = caption_row + 1
        for i, item in enumerate(block_mapping):
            block = item.get("block", "")
            ids = item.get("tasks", "")
            r = start_row + i
            ws.row_dimensions[r].height = 22

            c1 = ws.cell(row=r, column=1, value=block)
            c1.font = Font(bold=True, color="1F3864", size=10, name="Meiryo UI")
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.fill = PatternFill("solid", fgColor="D6E4F0")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

            c2 = ws.cell(row=r, column=3, value=ids)
            c2.font = Font(color="000000", size=10, name="Meiryo UI")
            c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)

        # adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12


# ─────────────────────────────────────────────
# Sheet 4: plain-language the roadmap
# ─────────────────────────────────────────────
def enhance_roadmap_sheet(ws, terminology: dict):
    """Plain-language the wording on the roadmap sheet."""
    # title
    roadmap_title = terminology.get("roadmap_title")
    if roadmap_title:
        ws.cell(row=1, column=1).value = roadmap_title

    # column headers
    header_map = terminology.get("roadmap_headers", {})
    for ci in range(1, 6):
        cell = ws.cell(row=2, column=ci)
        if cell.value in header_map:
            cell.value = header_map[cell.value]

    # rewrite the phase names
    phase_label_map = terminology.get("phase_labels", {})
    for ri in range(3, 6):
        cell = ws.cell(row=ri, column=1)
        if cell.value in phase_label_map:
            cell.value = phase_label_map[cell.value]

    # rewrite the 'watch out for' column (column 5)
    caution_map = terminology.get("caution_phrases", {})
    for ri in range(3, 6):
        cell = ws.cell(row=ri, column=5)
        if cell.value in caution_map:
            cell.value = caution_map[cell.value]
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # replace the wording in the benefit column (column 4)
    effect_map = terminology.get("effect_phrases", {})
    for ri in range(3, 6):
        cell = ws.cell(row=ri, column=4)
        if cell.value:
            v = cell.value
            for old, new in effect_map.items():
                v = v.replace(old, new)
            cell.value = v
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # soften the wording of the footnote row (the Human-in-the-loop principle)
    governance_note = terminology.get("governance_note")
    if governance_note:
        for ri in range(6, 12):
            cell = ws.cell(row=ri, column=1)
            if cell.value and "Human-in-the-loop" in str(cell.value):
                cell.value = governance_note
                cell.font = Font(italic=True, color="595959", size=9, name="Meiryo UI")
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)


# ─────────────────────────────────────────────
# automatic terminology replacement (jargon inside description fields)
# ─────────────────────────────────────────────
def substitute_terminology_in_cells(ws, terminology: dict):
    """Replace jargon with plain expressions across every string cell."""
    substitutions = terminology.get("terminology_substitutions", {})
    # skip the _comment field
    substitutions = {k: v for k, v in substitutions.items()
                     if not k.startswith("_")}
    if not substitutions:
        return

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value
                changed = False
                for old, new in substitutions.items():
                    if old in v:
                        v = v.replace(old, new)
                        changed = True
                if changed:
                    cell.value = v


# ─────────────────────────────────────────────
# main
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Enhance the AI readiness report (audience rewording, DAG swap, mapping table)"
    )
    parser.add_argument("--input", required=True,
                        help="the Excel file generated by generate_report.py")
    parser.add_argument("--output", required=True,
                        help="output Excel file")
    parser.add_argument("--audience", choices=["executive", "practitioner"],
                        default="practitioner",
                        help="audience: executive=plain language for executives / practitioner=for practitioners")
    parser.add_argument("--custom-dag-image", default=None,
                        help="replacement DAG image (PNG). The exported SVG drawn in Step 3")
    parser.add_argument("--block-mapping", default=None,
                        help="JSON file for the work-block mapping table")
    parser.add_argument("--terminology", default=str(DEFAULT_TERMINOLOGY),
                        help=f"terminology JSON (default: {DEFAULT_TERMINOLOGY})")
    args = parser.parse_args()

    # validate input
    if not os.path.exists(args.input):
        print(f"[error] Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    # copy, then edit
    shutil.copy(args.input, args.output)
    wb = openpyxl.load_workbook(args.output)

    # load the terminology dictionary (applied only in executive mode)
    terminology = {}
    if args.audience == "executive":
        terminology = load_terminology(Path(args.terminology))
        print(f"[1/4] Loaded terminology: {args.terminology}")

    # load the block mapping table
    block_mapping = []
    if args.block_mapping and os.path.exists(args.block_mapping):
        with open(args.block_mapping, encoding="utf-8") as f:
            block_mapping = json.load(f)
        print(f"[2/4] Loaded block mapping: {args.block_mapping} ({len(block_mapping)} entries)")
    else:
        print("[2/4] Block mapping: none (skipped)")

    # per-sheet processing
    sheet_names = wb.sheetnames

    # Sheet 1: summary
    if sheet_names:
        ws1 = wb.worksheets[0]
        if terminology:
            enhance_summary_sheet(ws1, terminology)
            substitute_terminology_in_cells(ws1, terminology)

    # Sheet 2: task list
    task_sheet_candidates = [s for s in sheet_names if "Task" in s]
    if task_sheet_candidates:
        ws2 = wb[task_sheet_candidates[0]]
        if terminology:
            enhance_task_sheet(ws2, terminology)
            substitute_terminology_in_cells(ws2, terminology)

    # Sheet 3: DAG
    dag_sheet_candidates = [s for s in sheet_names if "DAG" in s or "Dependency" in s]
    if dag_sheet_candidates:
        ws3 = wb[dag_sheet_candidates[0]]
        subtitle = terminology.get("dag_subtitle") if terminology else None
        if args.custom_dag_image or block_mapping:
            replace_dag_image(ws3, args.custom_dag_image, block_mapping, subtitle)
            print(f"[3/4] DAG sheet updated: image swapped={bool(args.custom_dag_image)}, "
                  f"mapping table added={bool(block_mapping)}")

    # Sheet 4: roadmap
    roadmap_sheet_candidates = [s for s in sheet_names
                                if "Roadmap" in s or "roadmap" in s]
    if roadmap_sheet_candidates:
        ws4 = wb[roadmap_sheet_candidates[0]]
        if terminology:
            enhance_roadmap_sheet(ws4, terminology)
            substitute_terminology_in_cells(ws4, terminology)

    wb.save(args.output)
    print(f"[4/4] Saved: {args.output}")
    print(f"\nDone. audience={args.audience}")


if __name__ == "__main__":
    main()
